"""Catalog sources: Quarterly Price List, Special Items, and the official
Allocated & Limited Distribution List xlsx.

Price list / Special Items (live-verified 2026-07-21):
    POST https://abc2.nc.gov/Pricing/PriceList   body: NCCode=&BrandName=<term>
    GET  https://abc2.nc.gov/Pricing/SpecialItems
    Row schema: NC Code | Supplier | Brand Name | Age | Proof | Size |
                Retail Price | MXB Price
    NC Codes here are formatted "18-650" (dash); the stock report uses
    "18650". normalize_nc_code() folds them together.
    NOTE: pricing data appears ~1 month before its effective date
    (pricing.abc.nc.gov schedule) — new codes here are an early signal.

Allocated xlsx (parsed live 2026-07-21):
    GET https://www.abc.nc.gov/local-abc-boards/public-allocated-and-limited-distribution-list/open
    Raw .xlsx at a stable URL, replaced in place. Contents: two columns
    (NC Code, Product), a label row like "Updated 1/1/2026", and section
    rows "ALLOCATED ITEMS" / "LIMITED DISTRIBUTION ITEMS" (~240 items).
    Diff by sha256 of the bytes; the landing page's "Last Updated" field is
    stale — trust the file, not the page.
"""
from __future__ import annotations

import hashlib
import io
import logging
import re
from dataclasses import dataclass

import openpyxl
from bs4 import BeautifulSoup

from ..http import fetch

log = logging.getLogger(__name__)

PRICELIST_URL = "https://abc2.nc.gov/Pricing/PriceList"
SPECIAL_ITEMS_URL = "https://abc2.nc.gov/Pricing/SpecialItems"
NEW_ITEMS_URL = "https://abc2.nc.gov/Pricing/ItemReports/2"
ALLOCATED_XLSX_URL = (
    "https://www.abc.nc.gov/local-abc-boards/"
    "public-allocated-and-limited-distribution-list/open"
)

PRICE_HEADERS = ["NC Code", "Supplier", "Brand Name", "Age", "Proof", "Size", "Retail Price", "MXB Price"]


@dataclass
class CatalogItem:
    nc_code: str
    brand_name: str
    retail_price: str
    source: str


@dataclass
class AllocatedItem:
    nc_code: str
    product: str
    section: str  # "ALLOCATED" or "LIMITED"


def normalize_nc_code(code: str) -> str:
    """'18-650' -> '18650'; strips whitespace. Stock report and Wake PLUs use
    the dashless form."""
    return re.sub(r"[^0-9]", "", code or "")


def _parse_price_tables(html: str, source: str) -> list[CatalogItem]:
    soup = BeautifulSoup(html, "lxml")
    items: list[CatalogItem] = []
    for table in soup.find_all("table"):
        first = table.find("tr")
        if not first:
            continue
        headers = [c.get_text(strip=True) for c in first.find_all(["th", "td"])]
        if "NC Code" not in headers or "Retail Price" not in headers:
            continue
        idx = {h: i for i, h in enumerate(headers)}
        for tr in table.find_all("tr")[1:]:
            cells = [td.get_text(strip=True) for td in tr.find_all("td")]
            if len(cells) < len(headers):
                continue  # category section rows etc.
            code = normalize_nc_code(cells[idx["NC Code"]])
            if not code:
                continue
            items.append(
                CatalogItem(
                    nc_code=code,
                    brand_name=cells[idx.get("Brand Name", 2)],
                    retail_price=cells[idx.get("Retail Price", 6)],
                    source=source,
                )
            )
    return items


def fetch_special_items(session, timeout: int = 60) -> list[CatalogItem]:
    resp = fetch(session, "GET", SPECIAL_ITEMS_URL, timeout=timeout)
    return _parse_price_tables(resp.text, "special_items")


def fetch_price_list(session, term: str, timeout: int = 60) -> list[CatalogItem]:
    resp = fetch(session, "POST", PRICELIST_URL, data={"NCCode": "", "BrandName": term}, timeout=timeout)
    return _parse_price_tables(resp.text, "price_list")


def fetch_new_items(session, timeout: int = 60) -> list[CatalogItem]:
    resp = fetch(session, "GET", NEW_ITEMS_URL, timeout=timeout)
    return _parse_price_tables(resp.text, "new_items")


def fetch_allocated_xlsx(session, timeout: int = 60) -> tuple[bytes, str]:
    resp = fetch(session, "GET", ALLOCATED_XLSX_URL, timeout=timeout)
    content = resp.content
    return content, hashlib.sha256(content).hexdigest()


def parse_allocated_xlsx(content: bytes) -> tuple[str, list[AllocatedItem]]:
    """Returns (label, items). Label is e.g. 'Updated 1/1/2026'."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    label = ""
    section = ""
    items: list[AllocatedItem] = []
    for row in ws.iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] is not None else ""
        b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        joined = (a + " " + b).strip()
        if not joined:
            continue
        if joined.lower().startswith("updated"):
            label = joined
            continue
        if "ALLOCATED ITEMS" in joined.upper():
            section = "ALLOCATED"
            continue
        if "LIMITED DISTRIBUTION" in joined.upper():
            section = "LIMITED"
            continue
        if a.lower() == "nc code":
            continue
        code = normalize_nc_code(a)
        if code and b:
            items.append(AllocatedItem(nc_code=code, product=b, section=section or "ALLOCATED"))
    return label, items
